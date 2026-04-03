from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


BASE_DIR = get_base_dir()
TEMPLATE_DIR = BASE_DIR / "ШАБЛОН_НЕ_ТРОГАТЬ"
TEMPLATE_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = TEMPLATE_DIR / "feedback_template.xlsx"


wb = Workbook()

ws_reviews = wb.active
ws_reviews.title = "Отзывы"
ws_dishes = wb.create_sheet("Блюда")
ws_reference = wb.create_sheet("Справочники")
ws_summary = wb.create_sheet("Сводка")
ws_capa = wb.create_sheet("CAPA")

tonality_list = ["Позитив", "Негатив", "Смешанный", "Не применимо"]
review_tag_list = [
    "Кухня",
    "Бар",
    "Сервис",
    "Зал",
    "Другое",
]
type_list = [
    "Бар / Бар",
    "Кухня / Кухня",
    "Кухня / Отравление",
    "Кухня / Включение",
    "Сервис / Встреча гостя",
    "Сервис / Коммуникация по телефону",
    "Сервис / Обслуживание",
    "Сервис / Принятие заказа",
    "Сервис / Расчёт + прощание",
    "Зал / Чистота",
    "Зал / АИ и игровая",
    "Зал / Атмосфера",
    "Зал / Прикассовая зона",
    "Зал / Безопасность",
    "Доставка / Кухня",
    "Доставка / Включение",
    "Доставка / Отравление",
    "Доставка / Время доставки",
    "Доставка / Не учли комментарий",
    "Доставка / Нет части заказа",
    "Доставка / Приборы",
    "Доставка / Упаковка",
    "Праздник / Кухня",
    "Праздник / Анимация",
    "Праздник / Организация",
    "Праздник / Сервис",
    "Праздник / Заказные торты",
    "Маркетинг / Программа лояльности",
    "Маркетинг / Афиша",
    "Маркетинг / Цена",
    "Маркетинг / Меню",
    "Маркетинг / Сайт",
    "Маркетинг / ЦП",
]
priority_list = ["Низкий", "Средний", "Высокий", "Критический"]
status_list = ["Запланировано", "В работе", "Закрыто", "Наблюдение"]
source_list = ["TableVisit", "GastroReview"]
problem_list = [
    "Не доложили / не доставили часть заказа",
    "Нет салфеток / приборов",
    "Блюдо невкусное",
    "Блюдо сырое / недоготовлено",
    "Долгое обслуживание",
    "Проблема с бонусами / промо",
    "Коммуникация / дезинформация",
    "Проблема организации / брони",
    "Вода / напиток плохого качества",
    "Проблема не определена",
]

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
input_fill = PatternFill("solid", fgColor="F8FBFF")
service_fill = PatternFill("solid", fgColor="EEF3F8")

negative_fill = PatternFill("solid", fgColor="F4A7A7")
mixed_fill = PatternFill("solid", fgColor="FFD966")
positive_fill = PatternFill("solid", fgColor="A9D18E")
priority_high_fill = PatternFill("solid", fgColor="FF8C00")
priority_critical_fill = PatternFill("solid", fgColor="FF4D4D")
overdue_fill = PatternFill("solid", fgColor="FF6B6B")


def style_sheet_grid(
    ws, rows, cols, center_cols=None, service_cols=None, date_cols=None
):
    center_cols = center_cols or []
    service_cols = service_cols or []
    date_cols = date_cols or []

    for row in range(2, rows + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in service_cols:
                cell.fill = service_fill
            else:
                cell.fill = input_fill
            if col in center_cols:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    for row in range(2, rows + 1):
        for col_letter in date_cols:
            ws[f"{col_letter}{row}"].number_format = "DD.MM.YYYY"


ws_reference["A1"] = "Тональность"
for i, item in enumerate(tonality_list, start=2):
    ws_reference[f"A{i}"] = item

ws_reference["B1"] = "Тип"
for i, item in enumerate(type_list, start=2):
    ws_reference[f"B{i}"] = item

ws_reference["C1"] = "Приоритет"
for i, item in enumerate(priority_list, start=2):
    ws_reference[f"C{i}"] = item

ws_reference["D1"] = "Статус"
for i, item in enumerate(status_list, start=2):
    ws_reference[f"D{i}"] = item

ws_reference["E1"] = "Источник"
for i, item in enumerate(source_list, start=2):
    ws_reference[f"E{i}"] = item

ws_reference["F1"] = "Проблематика"
for i, item in enumerate(problem_list, start=2):
    ws_reference[f"F{i}"] = item

ws_reference["G1"] = "Тег"
for i, item in enumerate(review_tag_list, start=2):
    ws_reference[f"G{i}"] = item

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_reference.column_dimensions[col].width = 36

for cell in ws_reference[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

ws_reference.sheet_state = "hidden"


# Лист Отзывы
review_headers = [
    "Дата",
    "Источник",
    "Кафе",
    "Стол",
    "Блюдо/упоминания",
    "Проблематика",
    "Суть отзыва",
    "Тональность",
    "Тип",
    "Приоритет",
    "Что сделали на месте",
    "Тег отзыва",
    "CAPA/Решение",
    "Ответственный",
    "Срок",
    "Статус",
    "Проверка результата",
    "Комментарий",
]

for col_num, header in enumerate(review_headers, start=1):
    ws_reviews.cell(row=1, column=col_num).value = header

for cell in ws_reviews[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

ws_reviews.row_dimensions[1].height = 30

review_widths = {
    "A": 14,
    "B": 16,
    "C": 24,
    "D": 10,
    "E": 28,
    "F": 28,
    "G": 50,
    "H": 14,
    "I": 30,
    "J": 14,
    "K": 28,
    "L": 18,
    "M": 28,
    "N": 18,
    "O": 14,
    "P": 16,
    "Q": 24,
    "R": 26,
}
for col, width in review_widths.items():
    ws_reviews.column_dimensions[col].width = width

style_sheet_grid(
    ws_reviews,
    rows=3000,
    cols=len(review_headers),
    center_cols=[1, 2, 4, 8, 10, 12, 14, 15, 16],
    service_cols=[2, 6, 8, 9, 10, 12, 15, 16],
    date_cols=["A", "O"],
)

ws_reviews.auto_filter.ref = "A1:R3000"
ws_reviews.freeze_panes = "A2"

review_table = Table(displayName="FeedbackTable", ref="A1:R3000")
review_table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
review_table.tableStyleInfo = review_table_style
ws_reviews.add_table(review_table)

dv_source = DataValidation(
    type="list", formula1="=Справочники!$E$2:$E$3", allow_blank=True
)
dv_problem = DataValidation(
    type="list",
    formula1=f"=Справочники!$F$2:$F${len(problem_list)+1}",
    allow_blank=True,
)
dv_tonality = DataValidation(
    type="list", formula1="=Справочники!$A$2:$A$4", allow_blank=True
)
dv_type = DataValidation(
    type="list", formula1=f"=Справочники!$B$2:$B${len(type_list)+1}", allow_blank=True
)
dv_priority = DataValidation(
    type="list", formula1="=Справочники!$C$2:$C$5", allow_blank=True
)
dv_status = DataValidation(
    type="list", formula1="=Справочники!$D$2:$D$5", allow_blank=True
)
dv_review_tag = DataValidation(
    type="list",
    formula1=f"=Справочники!$G$2:$G${len(review_tag_list)+1}",
    allow_blank=True,
)

for dv in [
    dv_source,
    dv_problem,
    dv_tonality,
    dv_type,
    dv_priority,
    dv_status,
    dv_review_tag,
]:
    ws_reviews.add_data_validation(dv)

dv_source.add("B2:B3000")
dv_problem.add("F2:F3000")
dv_tonality.add("H2:H3000")
dv_type.add("I2:I3000")
dv_priority.add("J2:J3000")
dv_review_tag.add("L2:L3000")
dv_status.add("P2:P3000")

for rng in ["A2:I3000", "K2:R3000"]:
    ws_reviews.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Негатив"'], fill=negative_fill)
    )
    ws_reviews.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Смешанный"'], fill=mixed_fill)
    )
    ws_reviews.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Позитив"'], fill=positive_fill)
    )

ws_reviews.conditional_formatting.add(
    "J2:J3000", FormulaRule(formula=['$J2="Высокий"'], fill=priority_high_fill)
)
ws_reviews.conditional_formatting.add(
    "J2:J3000", FormulaRule(formula=['$J2="Критический"'], fill=priority_critical_fill)
)
ws_reviews.conditional_formatting.add(
    "A2:R3000",
    FormulaRule(formula=['AND($O2<TODAY(),$P2<>"Закрыто",$O2<>"")'], fill=overdue_fill),
)


# Лист Блюда
dish_headers = [
    "Дата",
    "Источник",
    "Кафе",
    "Стол",
    "Блюдо",
    "Тег блюда",
    "Контекст/суть",
    "Тональность",
    "Приоритет",
    "Проблематика",
    "Комментарий",
]

for col_num, header in enumerate(dish_headers, start=1):
    ws_dishes.cell(row=1, column=col_num).value = header

for cell in ws_dishes[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

ws_dishes.row_dimensions[1].height = 30

dish_widths = {
    "A": 14,
    "B": 16,
    "C": 24,
    "D": 10,
    "E": 30,
    "F": 18,
    "G": 50,
    "H": 14,
    "I": 14,
    "J": 28,
    "K": 24,
}
for col, width in dish_widths.items():
    ws_dishes.column_dimensions[col].width = width

style_sheet_grid(
    ws_dishes,
    rows=5000,
    cols=len(dish_headers),
    center_cols=[1, 2, 4, 8, 9],
    service_cols=[2, 6, 8, 9, 10],
    date_cols=["A"],
)

ws_dishes.auto_filter.ref = "A1:K5000"
ws_dishes.freeze_panes = "A2"

dish_table = Table(displayName="DishTable", ref="A1:K5000")
dish_table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
dish_table.tableStyleInfo = dish_table_style
ws_dishes.add_table(dish_table)

dv_dish_tonality = DataValidation(
    type="list", formula1="=Справочники!$A$2:$A$5", allow_blank=True
)
dv_dish_priority = DataValidation(
    type="list", formula1="=Справочники!$C$2:$C$5", allow_blank=True
)
dv_dish_problem = DataValidation(
    type="list",
    formula1=f"=Справочники!$F$2:$F${len(problem_list)+1}",
    allow_blank=True,
)
dv_dish_tag = DataValidation(
    type="list",
    formula1=f"=Справочники!$G$2:$G${len(review_tag_list)+1}",
    allow_blank=True,
)

for dv in [dv_dish_tonality, dv_dish_priority, dv_dish_problem, dv_dish_tag]:
    ws_dishes.add_data_validation(dv)

dv_dish_tonality.add("H2:H5000")
dv_dish_priority.add("I2:I5000")
dv_dish_problem.add("J2:J5000")
dv_dish_tag.add("F2:F5000")

for rng in ["A2:H5000", "J2:K5000"]:
    ws_dishes.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Негатив"'], fill=negative_fill)
    )
    ws_dishes.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Смешанный"'], fill=mixed_fill)
    )
    ws_dishes.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Позитив"'], fill=positive_fill)
    )
    ws_dishes.conditional_formatting.add(
        rng, FormulaRule(formula=['$H2="Не применимо"'], fill=positive_fill)
    )

ws_dishes.conditional_formatting.add(
    "I2:I5000", FormulaRule(formula=['$I2="Высокий"'], fill=priority_high_fill)
)
ws_dishes.conditional_formatting.add(
    "I2:I5000", FormulaRule(formula=['$I2="Критический"'], fill=priority_critical_fill)
)
ws_dishes.conditional_formatting.add(
    "I2:I5000", FormulaRule(formula=['$I2="Низкий"'], fill=positive_fill)
)
ws_dishes.conditional_formatting.add(
    "I2:I5000", FormulaRule(formula=['$I2="Средний"'], fill=mixed_fill)
)


# Лист CAPA
ws_capa["A1"] = "CAPA реестр"
ws_capa["A1"].font = Font(size=14, bold=True)

capa_headers = [
    "ID",
    "Дата",
    "Источник",
    "Стол",
    "Блюдо/категория",
    "Проблематика",
    "Суть отзыва",
    "Тип",
    "Приоритет",
    "Решение/CAPA",
    "Ответственный",
    "Срок",
    "Статус",
    "Проверка результата",
    "Комментарий",
]

for col_num, header in enumerate(capa_headers, start=1):
    cell = ws_capa.cell(row=3, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

capa_widths = {
    "A": 10,
    "B": 14,
    "C": 16,
    "D": 10,
    "E": 28,
    "F": 28,
    "G": 46,
    "H": 26,
    "I": 14,
    "J": 28,
    "K": 18,
    "L": 14,
    "M": 16,
    "N": 22,
    "O": 24,
}
for col, width in capa_widths.items():
    ws_capa.column_dimensions[col].width = width

ws_capa.freeze_panes = "A4"
ws_capa.auto_filter.ref = "A3:O3000"

for row in range(4, 3001):
    for col in range(1, 16):
        cell = ws_capa.cell(row=row, column=col)
        cell.border = thin_border
        if col in [3, 6, 8, 9, 12, 13]:
            cell.fill = service_fill
        else:
            cell.fill = input_fill
        if col in [1, 2, 3, 4, 9, 11, 12, 13]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment

for row in range(4, 3001):
    ws_capa[f"B{row}"].number_format = "DD.MM.YYYY"
    ws_capa[f"L{row}"].number_format = "DD.MM.YYYY"

dv_capa_status = DataValidation(
    type="list", formula1="=Справочники!$D$2:$D$5", allow_blank=True
)
dv_capa_priority = DataValidation(
    type="list", formula1="=Справочники!$C$2:$C$5", allow_blank=True
)
dv_capa_type = DataValidation(
    type="list", formula1=f"=Справочники!$B$2:$B${len(type_list)+1}", allow_blank=True
)
dv_capa_problem = DataValidation(
    type="list",
    formula1=f"=Справочники!$F$2:$F${len(problem_list)+1}",
    allow_blank=True,
)

for dv in [dv_capa_status, dv_capa_priority, dv_capa_type, dv_capa_problem]:
    ws_capa.add_data_validation(dv)

dv_capa_problem.add("F4:F3000")
dv_capa_type.add("H4:H3000")
dv_capa_priority.add("I4:I3000")
dv_capa_status.add("M4:M3000")


# Легенда на Отзывах
ws_reviews["T1"] = "Легенда"
ws_reviews["T1"].font = Font(bold=True, size=12)

ws_reviews["T3"] = "Позитив"
ws_reviews["T4"] = "Смешанный"
ws_reviews["T5"] = "Негатив"
ws_reviews["T6"] = "Высокий приоритет"
ws_reviews["T7"] = "Критический приоритет"
ws_reviews["T8"] = "Просрочено"

ws_reviews["U3"].fill = positive_fill
ws_reviews["U4"].fill = mixed_fill
ws_reviews["U5"].fill = negative_fill
ws_reviews["U6"].fill = priority_high_fill
ws_reviews["U7"].fill = priority_critical_fill
ws_reviews["U8"].fill = overdue_fill

for row in range(3, 9):
    ws_reviews[f"T{row}"].border = thin_border
    ws_reviews[f"U{row}"].border = thin_border
    ws_reviews[f"T{row}"].alignment = left_alignment
    ws_reviews[f"U{row}"].alignment = center_alignment

ws_reviews.column_dimensions["T"].width = 24
ws_reviews.column_dimensions["U"].width = 14


# Лист Сводка
ws_summary["A1"] = "Сводка по отзывам и блюдам"
ws_summary["A1"].font = Font(size=14, bold=True)

ws_summary["A3"] = "Общая сводка"
ws_summary["A3"].font = Font(bold=True)

ws_summary["A4"] = "Всего отзывов"
ws_summary["B4"] = "=COUNTA(Отзывы!G2:G3000)"
ws_summary["A5"] = "Позитив"
ws_summary["B5"] = '=COUNTIF(Отзывы!H2:H3000,"Позитив")'
ws_summary["A6"] = "Смешанный"
ws_summary["B6"] = '=COUNTIF(Отзывы!H2:H3000,"Смешанный")'
ws_summary["A7"] = "Негатив"
ws_summary["B7"] = '=COUNTIF(Отзывы!H2:H3000,"Негатив")'
ws_summary["A8"] = "Просроченные"
ws_summary["B8"] = '=COUNTIFS(Отзывы!O2:O3000,"<"&TODAY(),Отзывы!P2:P3000,"<>Закрыто")'

for row in range(4, 9):
    ws_summary[f"A{row}"].border = thin_border
    ws_summary[f"B{row}"].border = thin_border
    ws_summary[f"A{row}"].alignment = left_alignment
    ws_summary[f"B{row}"].alignment = center_alignment

ws_summary["B5"].fill = positive_fill
ws_summary["B6"].fill = mixed_fill
ws_summary["B7"].fill = negative_fill
ws_summary["B8"].fill = overdue_fill


# Сводка по тегам отзывов
ws_summary["D3"] = "Сводка по тегам отзывов"
ws_summary["D3"].font = Font(bold=True)

ws_summary["D4"] = "Тег"
ws_summary["E4"] = "Позитив"
ws_summary["F4"] = "Смешанный"
ws_summary["G4"] = "Негатив"
ws_summary["H4"] = "Всего"

for cell in ws_summary["D4:H4"][0]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

for idx, tag in enumerate(review_tag_list, start=5):
    ws_summary[f"D{idx}"] = tag
    ws_summary[f"E{idx}"] = (
        f'=COUNTIFS(Отзывы!$L$2:$L$3000,$D{idx},Отзывы!$H$2:$H$3000,"Позитив")'
    )
    ws_summary[f"F{idx}"] = (
        f'=COUNTIFS(Отзывы!$L$2:$L$3000,$D{idx},Отзывы!$H$2:$H$3000,"Смешанный")'
    )
    ws_summary[f"G{idx}"] = (
        f'=COUNTIFS(Отзывы!$L$2:$L$3000,$D{idx},Отзывы!$H$2:$H$3000,"Негатив")'
    )
    ws_summary[f"H{idx}"] = f"=SUM(E{idx}:G{idx})"

    for col in ["D", "E", "F", "G", "H"]:
        ws_summary[f"{col}{idx}"].border = thin_border
        ws_summary[f"{col}{idx}"].alignment = (
            center_alignment if col != "D" else left_alignment
        )


# Сводка по тегам блюд
ws_summary["J3"] = "Сводка по тегам блюд"
ws_summary["J3"].font = Font(bold=True)

ws_summary["J4"] = "Тег блюда"
ws_summary["K4"] = "Позитив"
ws_summary["L4"] = "Смешанный"
ws_summary["M4"] = "Негатив"
ws_summary["N4"] = "Всего"

for cell in ws_summary["J4:N4"][0]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

for idx, tag in enumerate(review_tag_list, start=5):
    ws_summary[f"J{idx}"] = tag
    ws_summary[f"K{idx}"] = (
        f'=COUNTIFS(Блюда!$F$2:$F$5000,$J{idx},Блюда!$H$2:$H$5000,"Позитив")'
    )
    ws_summary[f"L{idx}"] = (
        f'=COUNTIFS(Блюда!$F$2:$F$5000,$J{idx},Блюда!$H$2:$H$5000,"Смешанный")'
    )
    ws_summary[f"M{idx}"] = (
        f'=COUNTIFS(Блюда!$F$2:$F$5000,$J{idx},Блюда!$H$2:$H$5000,"Негатив")'
    )
    ws_summary[f"N{idx}"] = f"=SUM(K{idx}:M{idx})"

    for col in ["J", "K", "L", "M", "N"]:
        ws_summary[f"{col}{idx}"].border = thin_border
        ws_summary[f"{col}{idx}"].alignment = (
            center_alignment if col != "J" else left_alignment
        )


# Сводка по блюдам
ws_summary["P3"] = "Сводка по блюдам"
ws_summary["P3"].font = Font(bold=True)

ws_summary["P4"] = "Блюдо"
ws_summary["Q4"] = "Тег блюда"
ws_summary["R4"] = "Позитив"
ws_summary["S4"] = "Смешанный"
ws_summary["T4"] = "Негатив"
ws_summary["U4"] = "Всего"

for cell in ws_summary["P4:U4"][0]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_alignment

# список блюд формулами не строим, потому что openpyxl не умеет удобно dynamic arrays.
# оставляем блок под ручную вставку или под второй этап.
# Но дадим готовые формулы для строк 5:300.
for row in range(5, 301):
    for col in ["P", "Q", "R", "S", "T", "U"]:
        ws_summary[f"{col}{row}"].border = thin_border
        ws_summary[f"{col}{row}"].alignment = (
            center_alignment if col != "P" else left_alignment
        )

    ws_summary[f"Q{row}"] = (
        f'=IF(P{row}="","",INDEX(Блюда!$F$2:$F$5000,MATCH(P{row},Блюда!$E$2:$E$5000,0)))'
    )
    ws_summary[f"R{row}"] = (
        f'=IF(P{row}="","",COUNTIFS(Блюда!$E$2:$E$5000,$P{row},Блюда!$H$2:$H$5000,"Позитив"))'
    )
    ws_summary[f"S{row}"] = (
        f'=IF(P{row}="","",COUNTIFS(Блюда!$E$2:$E$5000,$P{row},Блюда!$H$2:$H$5000,"Смешанный"))'
    )
    ws_summary[f"T{row}"] = (
        f'=IF(P{row}="","",COUNTIFS(Блюда!$E$2:$E$5000,$P{row},Блюда!$H$2:$H$5000,"Негатив"))'
    )
    ws_summary[f"U{row}"] = f'=IF(P{row}="","",SUM(R{row}:T{row}))'

ws_summary.column_dimensions["A"].width = 24
ws_summary.column_dimensions["B"].width = 14
ws_summary.column_dimensions["D"].width = 18
ws_summary.column_dimensions["E"].width = 12
ws_summary.column_dimensions["F"].width = 12
ws_summary.column_dimensions["G"].width = 12
ws_summary.column_dimensions["H"].width = 12
ws_summary.column_dimensions["J"].width = 18
ws_summary.column_dimensions["K"].width = 12
ws_summary.column_dimensions["L"].width = 12
ws_summary.column_dimensions["M"].width = 12
ws_summary.column_dimensions["N"].width = 12
ws_summary.column_dimensions["P"].width = 34
ws_summary.column_dimensions["Q"].width = 18
ws_summary.column_dimensions["R"].width = 12
ws_summary.column_dimensions["S"].width = 12
ws_summary.column_dimensions["T"].width = 12
ws_summary.column_dimensions["U"].width = 12

wb.save(OUTPUT_FILE)
print(f"Готово! Файл '{OUTPUT_FILE}' создан.")
